import tkinter
from tkinter import ttk
from tkinter import messagebox
from tkinter import *
import openpyxl
import os
from openpyxl.utils.exceptions import InvalidFileException
import datetime
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook

#global variables
custname = ''
deliverytime = ''
ordertype = ''
deliveryperson = ''
paidstatus = ''
calc_price = 0
global calc_count
calc_count  = 0
#row_inserted = False
#print("global", calc_count)
#function to calculate total price
def get_calc():
    
    #get details from GUI fields
    custname = cust_name_entry.get()
    deliverytime = deliverytime_entry.get()
    ordertype = ordertype_combobox.get()
    deliveryperson = deliveryperson_combobox.get()
    paidstatus = paid_status_var.get()
    halftiffin = halftiffin_var.get()
    halfbhaji1 = halfbhaji1_var.get()
    halfbhaji2 = halfbhaji2_var.get()
    halfvaran = halfvaran_var.get()
    halfrice = halfrice_var.get()

    #print(halftiffin, halfbhaji1, halfbhaji2, halfvaran, halfrice)
    #print("paid status", paidstatus)



    tiffin_quant = tiffin_spinbox.get()
    tiffin_quant = float(tiffin_quant)

    chapati_quant = chapati_spinbox.get()
    chapati_quant = float(chapati_quant)

    bhakari_quant = bhakari_spinbox.get()
    bhakari_quant = float(bhakari_quant)

    bhaji1_quant = bhaji1_spinbox.get()
    bhaji1_quant = float(bhaji1_quant)

    bhaji2_quant = bhaji2_spinbox.get()
    bhaji2_quant = float(bhaji2_quant)

    varan_quant = varan_spinbox.get()
    varan_quant = float(varan_quant)

    rice_quant = rice_spinbox.get()
    rice_quant = float(rice_quant)

    thepla_quant = thepla_spinbox.get()
    thepla_quant = float(thepla_quant)

    modak_quant = modak_spinbox.get()
    modak_quant = float(modak_quant)

    poli_quant = poli_spinbox.get()
    poli_quant = float(poli_quant)

    bhakri_type = bhakri_entry.get()
    bhaji1_type = bhaji1_entry.get()
    bhaji2_type = bhaji2_entry.get()

    #concatenating values for order details 
    order_details = (str(int(tiffin_quant)) + " Tiffin" + " | \n" + 
                     str(int(chapati_quant)) + " Chapati" + " | \n" + 
                     str(int(bhakari_quant)) + " " + str(bhakri_type) + " Bhakri" + " " + " | \n" +  
                     str(int(bhaji1_quant)) + " " + str(bhaji1_type) + " Bhaji 1" + " " + " | \n" + 
                     str(int(bhaji2_quant)) + " " + str(bhaji2_type) + " Bhaji 2" + " " + " | \n" + 
                     str(int(varan_quant)) + " Varan" + " | \n" + 
                     str(int(rice_quant)) + " Rice" + " | \n" )

    order_details_thepla = str(int(thepla_quant))
    order_details_modak = str(int(modak_quant))
    order_details_poli = str(int(poli_quant))


    #Get Prices from admin datasheet
    price_filepath = "ADMIN_PRICE.xlsx"
    
    try:

        wb = openpyxl.load_workbook(price_filepath)
        ws = wb.active
        #cell_value = ws.cell(row=1, column =1).value
        #print(cell_value)
        #calculate total price
        #consider half checkbox

        if halftiffin == 'Half':
            Tiffin_PerPlate = ws.cell(row=2, column =3).value
        else:
            Tiffin_PerPlate = ws.cell(row=2, column =2).value


        if halfbhaji1 == 'Half':
            Bhaji1_PerPlate = ws.cell(row=5, column =3).value
        else:
            Bhaji1_PerPlate = ws.cell(row=5, column =2).value


        if halfbhaji2 == 'Half':
            Bhaji2_PerPlate = ws.cell(row=6, column =3).value
        else:
            Bhaji2_PerPlate = ws.cell(row=6, column =2).value


        if halfvaran == 'Half':
            Varan_PerPlate = ws.cell(row=7, column =3).value
        else:
            Varan_PerPlate = ws.cell(row=7, column =2).value


        if halfrice == 'Half':
            Rice_PerPlate = ws.cell(row=8, column =3).value
        else:
            Rice_PerPlate = ws.cell(row=8, column =2).value


        Chapati_PerPlate = ws.cell(row=3, column =2).value
        Bhakri_PerPlate = ws.cell(row=4, column =2).value
        Thepla_PerPlate = ws.cell(row=9, column =2).value
        Modak_PerPlate = ws.cell(row=10, column =2).value
        Poli_PerPlate = ws.cell(row=11, column =2).value
    

        #print(Tiffin_PerPlate, Bhaji1_PerPlate, Bhaji2_PerPlate, Varan_PerPlate, Rice_PerPlate)

        #reset to zero for next calculation
        calc_price = 0.0
        price_entry = tkinter.Label(order_detail_frame, textvariable=calc_price,font=("Roboto", 20, "bold"), width=7)
        price_entry.grid(row=1,column=2)
        
        #formula to calculate total price of order
        calc_price =    (tiffin_quant * float(Tiffin_PerPlate)) + \
                        (chapati_quant * float(Chapati_PerPlate)) + \
                        (bhakari_quant * float(Bhakri_PerPlate)) + \
                        (bhaji1_quant * float(Bhaji1_PerPlate)) + \
                        (bhaji2_quant * float(Bhaji2_PerPlate)) + \
                        (varan_quant * float(Varan_PerPlate)) + \
                        (rice_quant * float(Rice_PerPlate)) + \
                        (thepla_quant * float(Thepla_PerPlate)) + \
                        (modak_quant * float(Modak_PerPlate)) + \
                        (poli_quant * float(Poli_PerPlate))
        
        #creating list to return to get_info function
        #global calc_count 
        global calc_count
        calc_count += 1
        #print("inside calc ",calc_count)
        list1 = [order_details, calc_price, order_details_thepla, order_details_modak, order_details_poli, custname, ordertype, calc_count,
                paidstatus, deliverytime, deliveryperson]

        #displaying calcuated value on GUI
        calc_price = tkinter.StringVar(value = calc_price)
        price_entry = tkinter.Label(order_detail_frame, textvariable=calc_price, font=("Roboto", 20, "bold"), width=7)
        price_entry.grid(row=1,column=2)

        wb.save(price_filepath)
    except:
        messagebox.showwarning(title="ERROR !", message="Please Close ADMIN PRICE Excel File")
    
    return list1


def Clear_Window():

    #clean window after submit button is clicked

    cust_name_entry.delete(0,tkinter.END)
    deliverytime_entry.delete(0,tkinter.END)
    bhaji1_entry.delete(0,tkinter.END)
    bhaji2_entry.delete(0,tkinter.END)
    bhakri_entry.delete(0,tkinter.END)

    calc_price = 0.0
    price_entry = tkinter.Label(order_detail_frame, textvariable=calc_price,font=("Roboto", 20, "bold"), width=7)
    price_entry.grid(row=1,column=2)

    ordertype_combobox.set('')
    deliveryperson_combobox.set('')
    amountpaid_check.deselect()
    halftiffin_check.deselect()
    halfbhaji1_check.deselect()
    halfbhaji2_check.deselect()
    halfvaran_check.deselect()
    halfrice_check.deselect()

    reset_v = IntVar(window)
    reset_v.set(0)
    tiffin_spinbox.config(textvariable = reset_v)
    reset_v1 = IntVar(window)
    reset_v1.set(0)
    chapati_spinbox.config(textvariable = reset_v1)
    reset_v2 = IntVar(window)
    reset_v2.set(0)
    bhakari_spinbox.config(textvariable = reset_v2)
    reset_v3 = IntVar(window)
    reset_v3.set(0)
    bhaji1_spinbox.config(textvariable = reset_v3)
    reset_v4 = IntVar(window)
    reset_v4.set(0)
    bhaji2_spinbox.config(textvariable = reset_v4)
    reset_v5 = IntVar(window)
    reset_v5.set(0)
    varan_spinbox.config(textvariable = reset_v5)
    reset_v6 = IntVar(window)
    reset_v6.set(0)
    rice_spinbox.config(textvariable = reset_v6)
    reset_v7 = IntVar(window)
    reset_v7.set(0)
    thepla_spinbox.config(textvariable = reset_v7)
    reset_v8 = IntVar(window)
    reset_v8.set(0)
    modak_spinbox.config(textvariable = reset_v8)
    reset_v9 = IntVar(window)
    reset_v9.set(0)
    poli_spinbox.config(textvariable = reset_v9)    

    print("inside clear window")

#function to save everything into excel
def get_info():

    #price = price_entry.get()
    #price = float(price)
    #price_f = "{:.2f}".format(price)

    #fetching values from get_calc function
    global calc_count
    list2 = get_calc()
    price_dec = list2[1]
    custname = list2[5]
    ordertype = list2[6]
    orderdetails = list2[0]
    paidstatus = list2[8]
    deliverytime = list2[9]
    deliveryperson = list2[10]

    blankrow_inserted = None
    #print(paidornot)
    #calc_count = list2[7]
    #print("flag inside info ", calc_count)
    
    #valition for mandatory fields > custname and ordertype
    if custname:
        if ordertype:
            if calc_count > 1:    
                #get date
                today = datetime.date.today() 
                #filepath = "C:\\Users\\mhatr\\OneDrive\\Tkinter\\AkshayPatra_DailyOrders_" + str(today) + ".xlsx"   
                filename = "AkshayPatra_DailyOrders_" + str(today) + ".xlsx"   
                user_dir = os.path.expanduser('~')
                filepath = os.path.join(user_dir, filename)
                
                #if file is not present then create new file
                if not os.path.exists(filepath):
                    workbook = openpyxl.Workbook()
                    sheet = workbook.active
                    heading = ["Customer Name", "Delivery Time", "Price", "Order Type", 
                                "Payment Status", "Delivery Person", "Order Details", "Thepla", "Modak", "Poli"]
                    sheet.append(heading)
                    
                    #color the header
                    for rows in sheet.iter_rows(min_row=1, max_row=1, min_col=1):
                        for cell in rows:
                            cell.fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

                    order_no = 1
                    workbook.save(filepath)
                
                try: 
                    workbook = openpyxl.load_workbook(filepath)
                    sheet = workbook.active
                    
                    #separate morning and evening orders
                    # Check if the current time is after 2 pm
                    print("inside get info")
                    current_time = datetime.datetime.now().time()
                    print(current_time)
                    print(datetime.time(hour=15))
                    
                    if current_time >= datetime.time(hour=1):
                        print("it's par 3 Oclk")
                        # Check if a blank row has already been inserted
                        blankrow_inserted = False
                        for row in sheet.iter_rows():
                            if all(cell.value is None for cell in row):
                                blankrow_inserted = True
                                break

                    
                    print("outside")
                    #print(row_inserted) 
                    # If a blank row has not been inserted, insert one
                    if blankrow_inserted == False:
                        blank_list = []
                        print("blank row inserted")
                        sheet.append(blank_list)
                        sheet.append(blank_list)
                        #print("blank row inserted")
                    
                    if not deliverytime:
                        deliverytime = current_time.strftime('%I:%M %p')

                    sheet.append([custname, deliverytime, str(price_dec), ordertype, paidstatus, deliveryperson, orderdetails, list2[2], list2[3], list2[4]])
                    
                    #highlight paid transactions
                    #bold_font = Font(bold=True)
                    fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                    
                    for row in sheet.iter_rows(min_row=2): # skip the header row
                        #print("paidstatus value", row[4].value)
                        if row[4].value == "Paid": # column 4 is the "paidstatus" column
                            for cell in row:
                                cell.fill = fill

                    print("inside submit2")
                    workbook.save(filepath)
                    #reset counter
                    calc_count = 0
                    #print("ending", calc_count)
                    Clear_Window()

                except:
                    tkinter.messagebox.showwarning(title="ERROR !", message="Please Close Akshaypatra Excel File")
            else:
                tkinter.messagebox.showwarning(title="ERROR !", message="Calculate first")
        else: 
            tkinter.messagebox.showwarning(title="ERROR !", message="Order Type missing")
    else:
        tkinter.messagebox.showwarning(title="ERROR !", message="Customer Name missing")



#Tkinter GUI code

window = tkinter.Tk()
window.title("Order Entry Form")
frame = tkinter.Frame(window)
frame.pack()

reset_v = IntVar(window)
reset_v.set(0)

#frame 01: order details
order_detail_frame = tkinter.LabelFrame(frame, text="Order Details")
order_detail_frame.grid(row=0, column=0, padx=20, pady=10)

#save customer details
cust_name_label = tkinter.Label(order_detail_frame, text='Customer Name')
cust_name_label.grid(row=0,column=0)
cust_name_entry = tkinter.Entry(order_detail_frame)
cust_name_entry.grid(row=1,column=0)

#save item details
deliverytime_label = tkinter.Label(order_detail_frame, text="Delivery Time")
deliverytime_label.grid(row=0, column=1)
deliverytime_entry = tkinter.Entry(order_detail_frame)
#deliverytime_entry = ttk.TimeEntry(order_detail_frame, width=12, format='HH:mm:ss')
deliverytime_entry.grid(row=1,column=1)

#save total order price
price_label = tkinter.Label(order_detail_frame, text="Total Price")
price_label.grid(row=0, column=2)
calc_price = tkinter.StringVar(value = 0.0)
price_entry = tkinter.Label(order_detail_frame, textvariable=calc_price,font=("Roboto", 20, "bold"), width=7)
price_entry.grid(row=1,column=2)

#check if delivery or takeaway
ordertype_label = tkinter.Label(order_detail_frame, text="Order Type")
ordertype_label.grid(row=2,column=0)
ordertype_combobox = ttk.Combobox(order_detail_frame, values=["Take Away","Delivery"])
ordertype_combobox.grid(row=3,column=0)

#calculate button
calc_button = tkinter.Button(order_detail_frame, text="Calculate Price", command= get_calc)
calc_button.grid(row=3, column=1, padx=20, pady=20)

#adding padding for all the widgets inside frame 
for widget in order_detail_frame.winfo_children():
    widget.grid_configure(padx=10, pady=5)


#frame 02: delivery details
delivey_detail_frame = tkinter.LabelFrame(frame, text="Delivery Details")
delivey_detail_frame.grid(row=1, column=0, sticky="news", padx=20, pady=20)

#select delivery person
deliveryperson_label = tkinter.Label(delivey_detail_frame, text="Delivery Person")
deliveryperson_label.grid(row=0,column=0)
deliveryperson_combobox = ttk.Combobox(delivey_detail_frame, values=["Lomesh","Rajas"])
deliveryperson_combobox.grid(row=1,column=0)

#check if amount is paid or not
amountpaid_label = tkinter.Label(delivey_detail_frame, text="Payment Status")
amountpaid_label.grid(row=0, column=1)

paid_status_var = tkinter.StringVar(value="Not Paid")
amountpaid_check = tkinter.Checkbutton(delivey_detail_frame, text="Paid", 
                                       variable=paid_status_var, onvalue="Paid", offvalue="Not Paid")
amountpaid_check.grid(row=1, column=1)

#adding padding for all the widgets inside frame 
for widget in delivey_detail_frame.winfo_children():
    widget.grid_configure(padx=10, pady=5)



#frame 03: Submit it
menu_detail_frame = tkinter.LabelFrame(frame, text="Menu")
menu_detail_frame.grid(row=2, column=0, sticky="news", padx=20, pady=10)

#menu items grid
quantity_label = tkinter.Label(menu_detail_frame, text="Quantity")
quantity_label.grid(row=0, column=1)

tiffin_label = tkinter.Label(menu_detail_frame, text="Tiffin")
tiffin_label.grid(row=1, column=0)
tiffin_spinbox = tkinter.Spinbox(menu_detail_frame, from_=0, to=99)
tiffin_spinbox.grid(row=1,column=1)

chapati_label = tkinter.Label(menu_detail_frame, text="Chapati")
chapati_label.grid(row=2, column=0)
chapati_spinbox = tkinter.Spinbox(menu_detail_frame, from_=0, to=99)
chapati_spinbox.grid(row=2,column=1)

bhakari_label = tkinter.Label(menu_detail_frame, text="Bhakari")
bhakari_label.grid(row=3, column=0)
bhakari_spinbox = tkinter.Spinbox(menu_detail_frame, from_=0, to=99)
bhakari_spinbox.grid(row=3,column=1)

bhaji1_label = tkinter.Label(menu_detail_frame, text="Bhaji 1")
bhaji1_label.grid(row=4, column=0)
bhaji1_spinbox = tkinter.Spinbox(menu_detail_frame, from_=0, to=99)
bhaji1_spinbox.grid(row=4,column=1)

bhaji2_label = tkinter.Label(menu_detail_frame, text="Bhaji 2")
bhaji2_label.grid(row=5, column=0)
bhaji2_spinbox = tkinter.Spinbox(menu_detail_frame, from_=0, to=99)
bhaji2_spinbox.grid(row=5,column=1)

varan_label = tkinter.Label(menu_detail_frame, text="Varan")
varan_label.grid(row=6, column=0)
varan_spinbox = tkinter.Spinbox(menu_detail_frame, from_=0, to=99)
varan_spinbox.grid(row=6,column=1)

rice_label = tkinter.Label(menu_detail_frame, text="Rice")
rice_label.grid(row=7, column=0)
rice_spinbox = tkinter.Spinbox(menu_detail_frame, from_=0, to=99)
rice_spinbox.grid(row=7,column=1)

thepla_label = tkinter.Label(menu_detail_frame, text="Thepla")
thepla_label.grid(row=8, column=0)
thepla_spinbox = tkinter.Spinbox(menu_detail_frame, from_=0, to=99)
thepla_spinbox.grid(row=8,column=1)

modak_label = tkinter.Label(menu_detail_frame, text="Modak")
modak_label.grid(row=9, column=0)
modak_spinbox = tkinter.Spinbox(menu_detail_frame, from_=0, to=99)
modak_spinbox.grid(row=9,column=1)

poli_label = tkinter.Label(menu_detail_frame, text="Puran Poli")
poli_label.grid(row=10, column=0)
poli_spinbox = tkinter.Spinbox(menu_detail_frame, from_=0, to=99)
poli_spinbox.grid(row=10,column=1)



#half prices
type_label = tkinter.Label(menu_detail_frame, text="Half")
type_label.grid(row=0, column=2)

halftiffin_var = tkinter.StringVar(value="Full")
halftiffin_check = tkinter.Checkbutton(menu_detail_frame, text="", variable=halftiffin_var, onvalue="Half", offvalue="Full")
halftiffin_check.grid(row=1, column=2)

halfbhaji1_var = tkinter.StringVar(value="Full")
halfbhaji1_check = tkinter.Checkbutton(menu_detail_frame, text="", variable=halfbhaji1_var, onvalue="Half", offvalue="Full")
halfbhaji1_check.grid(row=4, column=2)

halfbhaji2_var = tkinter.StringVar(value="Full")
halfbhaji2_check = tkinter.Checkbutton(menu_detail_frame, text="", variable=halfbhaji2_var, onvalue="Half", offvalue="Full")
halfbhaji2_check.grid(row=5, column=2)

halfvaran_var = tkinter.StringVar(value="Full")
halfvaran_check = tkinter.Checkbutton(menu_detail_frame, text="", variable=halfvaran_var, onvalue="Half", offvalue="Full")
halfvaran_check.grid(row=6, column=2)

halfrice_var = tkinter.StringVar(value="Full")
halfrice_check = tkinter.Checkbutton(menu_detail_frame, text="", variable=halfrice_var, onvalue="Half", offvalue="Full")
halfrice_check.grid(row=7, column=2)



#types of menu
type_label = tkinter.Label(menu_detail_frame, text="Type")
type_label.grid(row=0, column=3)

bhakri_entry = tkinter.Entry(menu_detail_frame)
bhakri_entry.grid(row=3,column=3)

bhaji1_entry = tkinter.Entry(menu_detail_frame)
bhaji1_entry.grid(row=4,column=3)

bhaji2_entry = tkinter.Entry(menu_detail_frame)
bhaji2_entry.grid(row=5,column=3)

#adding padding for all the widgets inside frame 
for widget in menu_detail_frame.winfo_children():
    widget.grid_configure(padx=10, pady=5)


#submit button
button = tkinter.Button(frame, text="Submit", command= get_info)
button.grid(row=3, column=0, sticky="news", padx=20, pady=20)

#to keep GUI open until close by user
window.mainloop()

